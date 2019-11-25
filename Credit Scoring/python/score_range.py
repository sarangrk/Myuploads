# -*- coding: utf-8 -*-
"""
Created on Thu Sep  6 10:27:24 2018

@author: ie186001
"""

import os
import openpyxl
import logging
import re
#import yaml
import xlrd
import teradata

#open excel template for score distribution
scr_workbook='C:/Users/ie186001/Documents/CLIENT DATA/JapanDAP/Code/src/python/ScoreRangeMetadata.xlsm'

#wb_src = openpyxl.load_workbook(scr_workbook)
#ws_meta = wb_src.get_sheet_by_name("ScrRange")
#wb_src.close()

wb_src = xlrd.open_workbook(scr_workbook)
ws_meta = wb_src.sheet_by_name("ScrRange")

#get database input value from excel
db_val = ws_meta.cell(0,7).value

#get score cut-off input value from excel
cutoff = ws_meta.cell(0,4).value
cutoff = int(cutoff)
cutoff_t = (cutoff,)

#get values from excel file
values=[]
cnt=0

for r in range(2, ws_meta.nrows):
    l_bound	 = ws_meta.cell(r,0).value
    h_bound	 = ws_meta.cell(r,1).value

    x=(l_bound, h_bound,cnt)

    # Assign values from each row
    values.append(x)   

    #print(values)
    cnt+=1

#create Teradata connection
udaExec = teradata.UdaExec(appName="JapanTD", version="1.0", logConsole=False)
session = udaExec.connect(method="odbc", dsn="JapanDAP64",username="dbc", password="dbc", autocommit=True,transactionMode="Teradata")

#create delete and insert query
#score range
sql_delete_range="DELETE FROM %s.TBL_BASE_SCORE;" % db_val
sql_insert_range="INSERT INTO %s.tbl_base_score VALUES (?, ?,?)" % db_val

#score cut-off
sql_delete_cutoff="DELETE FROM %s.tbl_score_cutoff;" % db_val
sql_insert_cutoff="INSERT INTO %s.tbl_score_cutoff VALUES (?,current_timestamp(6))" % db_val

#execute queries
session.execute(sql_delete_range)
session.executemany(sql_insert_range, values,batch=True)

session.execute(sql_delete_cutoff)
session.execute(sql_insert_cutoff,cutoff_t)

session.close()






